from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet


class ExcelWriter:
    headers = [
        "filename",
        "FaceNet (%)",
        "LPIPS (%)",
        "Final score (%)",
        "FER emotion (%)",
        "DeepFace emotion (%)",
        "Person count",
        "Duration",
    ]

    def write(self, data: object, output_path: Path) -> None:
        records = self._normalize_records(data)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "metrics"
        sheet.append(self.headers)
        for record in records:
            sheet.append(
                [
                    record.get("filename", ""),
                    float(record.get("facenet_percent", 0.0)),
                    float(record.get("lpips_percent", 0.0)),
                    float(record.get("final_percent", 0.0)),
                    float(record.get("fer_percent", 0.0)),
                    float(record.get("deepface_percent", 0.0)),
                    int(record.get("person_count", 0)),
                    record.get("duration_label", ""),
                ]
            )
        self._format_sheet(sheet)
        workbook.save(output_path)

    def _normalize_records(self, data: object) -> List[Mapping[str, object]]:
        if hasattr(data, "to_dict"):
            return list(data.to_dict("records"))
        if isinstance(data, Sequence):
            return [record for record in data]
        if isinstance(data, Iterable):
            return list(data)
        return []

    def _format_sheet(self, sheet: Worksheet) -> None:
        alignment = Alignment(horizontal="center", vertical="center")
        bold_font = Font(bold=True)
        thick = Side(style="thick")
        max_col = len(self.headers)
        max_row = sheet.max_row
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = alignment
            row[3].font = bold_font
        for col in range(1, max_col + 1):
            self._apply_border(sheet.cell(row=1, column=col), top=thick, bottom=thick)
            self._apply_border(sheet.cell(row=max_row, column=col), bottom=thick)
        for row in range(1, max_row + 1):
            self._apply_border(sheet.cell(row=row, column=1), left=thick)
            self._apply_border(sheet.cell(row=row, column=max_col), right=thick)
            self._apply_border(sheet.cell(row=row, column=3), right=thick)
            self._apply_border(sheet.cell(row=row, column=4), left=thick, right=thick)
            self._apply_border(sheet.cell(row=row, column=5), left=thick)
        for row_index in range(2, max_row + 1):
            for col_index in (2, 3, 4, 5, 6):
                sheet.cell(row=row_index, column=col_index).number_format = "0.0"
            sheet.cell(row=row_index, column=7).number_format = "0"
        self._set_column_widths(sheet)

    def _apply_border(self, cell, left: Side | None = None, right: Side | None = None, top: Side | None = None, bottom: Side | None = None) -> None:
        cell.border = Border(
            left=left or cell.border.left,
            right=right or cell.border.right,
            top=top or cell.border.top,
            bottom=bottom or cell.border.bottom,
        )

    def _set_column_widths(self, sheet: Worksheet) -> None:
        longest = max((len(str(sheet.cell(row=i, column=1).value)) for i in range(1, sheet.max_row + 1)), default=0)
        sheet.column_dimensions["A"].width = max(30, longest + 5)
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 12
        sheet.column_dimensions["D"].width = 16
        sheet.column_dimensions["E"].width = 22
        sheet.column_dimensions["F"].width = 24
        sheet.column_dimensions["G"].width = 14
        sheet.column_dimensions["H"].width = 14
