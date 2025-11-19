from __future__ import annotations

import argparse
import logging
import random
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from openpyxl import Workbook
from tqdm import tqdm

from .config import PipelineConfig, require_cuda_device
from .pipeline import MetricsPipeline
from .sample_workbook import ColumnDefinition, SampleWorkbook, SheetSpec


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Face de-identification and emotion metrics pipeline")
    parser.add_argument("--base-dir", required=True, help="Base directory containing mirrored input and output folders")
    parser.add_argument("--output", help="Excel file to create (defaults to <base>/rapa_report_samples.xlsx)")
    parser.add_argument("--max-frames-per-video", type=int, default=16, help="Maximum frames sampled per video")
    parser.add_argument("--lpips-distance-max", type=float, default=1.0, help="Maximum LPIPS distance mapped to 0 percent similarity")
    parser.add_argument("--max-files", type=int, help="Maximum number of files to process for debugging")
    parser.add_argument("--video-backend", default="auto", help="Video decoding backend to use (auto tries decord then ffmpeg)")
    parser.add_argument("--profile-only", action="store_true", help="Process a 10-sample deterministic profile run instead of writing Excel")
    parser.add_argument("--seed", type=int, default=42, help="Seed used for sampling files in profiling mode")
    parser.add_argument("--resize-512", action="store_true", help="Resize each frame or image to 512x512 before inference")
    parser.add_argument("--profile-kind", choices=["all", "image", "video"], default="all", help="Restrict profiling samples to images, videos, or all")
    parser.add_argument("--debug-random-10", action="store_true", help="Generate a 10-row random debug sample (5 videos + 5 images)")
    parser.add_argument("--debug-output", default="random_debug_10.xlsx", help="Path for the debug sample workbook")
    parser.add_argument(
        "--top-bottom-40-only",
        action="store_true",
        help="Process only the first 20 and last 20 sorted paths (skip random sample_100/sample_500 sheets)",
    )
    parser.add_argument(
        "--random-sample",
        type=int,
        help="Process a random subset of matched files and write them to a standalone workbook",
    )
    parser.add_argument(
        "--random-output",
        help="Excel path for the random sample workbook (defaults to <base>/random_sample_<N>.xlsx)",
    )
    parser.add_argument("--log-level", default="INFO", help="Logging level")
    return parser.parse_args()


def _print_profile_summary(result) -> None:
    print(f"Profiling {result.processed} of {result.requested} requested pairs")
    totals = dict(result.totals)
    total_duration = totals.get("total", result.elapsed)
    stages = ["load", "resize", "facenet", "lpips", "emoti", "other", "total"]
    for stage in stages:
        duration = totals.get(stage, 0.0)
        milliseconds = duration * 1000.0
        if stage == "total":
            percent = 100.0 if total_duration > 0 else 0.0
        elif total_duration > 0:
            percent = (duration / total_duration) * 100.0
        else:
            percent = 0.0
        print(f"{stage:<7}: {milliseconds:9.0f} ms ({percent:5.1f}%)")


def _write_debug_sample(records, output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "debug_sample"
    headers = ["Filename", "Facenet", "LPIPS", "Final Score", "Emoti", "Person Count", "Duration"]
    sheet.append(headers)
    for record in records:
        sheet.append(
            [
                record.get("filename", ""),
                _safe_float(record.get("facenet_percent")),
                _safe_float(record.get("lpips_percent")),
                _safe_float(record.get("final_score_percent")),
                _safe_float(record.get("emoti_emotion_percent")),
                _safe_int(record.get("person_count")),
                record.get("duration_label", ""),
            ]
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _safe_float(value: object):
    if value in (None, ""):
        return ""
    try:
        return float(value)
    except Exception:
        return value


def _safe_int(value: object):
    if value in (None, ""):
        return ""
    try:
        return int(value)
    except Exception:
        return value


def _print_debug_sample(records) -> None:
    print("Random debug sample (filename | FaceNet | LPIPS | Final score)")

    def fmt(value):
        if value is None:
            return "NA"
        return f"{value:5.1f}"

    for record in records:
        facenet = record.get("facenet_percent")
        lpips = record.get("lpips_percent")
        final_score = record.get("final_score_percent")
        line = f"{record.get('filename',''):<20} | {fmt(facenet)} | {fmt(lpips)} | {fmt(final_score)}"
        print(line)


TABLE_HEADERS = ["Filename", "Facenet", "LPIPS", "Final Score", "Emoti", "Person Count", "Duration"]
DECIMAL_HEADERS = ["Facenet", "LPIPS", "Final Score", "Emoti"]
INTEGER_HEADERS = ["Person Count"]
THICK_BOUNDARIES = [("LPIPS", "Final Score"), ("Final Score", "Emoti")]
COLUMN_DEFINITIONS: Sequence[ColumnDefinition] = [
    ColumnDefinition("Filename", "filename"),
    ColumnDefinition("Facenet", "facenet_percent"),
    ColumnDefinition("LPIPS", "lpips_percent"),
    ColumnDefinition("Final Score", "final_score_percent"),
    ColumnDefinition("Emoti", "emoti_emotion_percent"),
    ColumnDefinition("Person Count", "person_count"),
    ColumnDefinition("Duration", "duration_label"),
]


def _default_output_path(base_dir: Path, override: str | None) -> Path:
    if override:
        return Path(override).expanduser()
    return (base_dir / "rapa_report_samples.xlsx").expanduser()


def _first_last(paths: List[str], count: int) -> List[str]:
    if not paths or count <= 0:
        return []
    count = min(count, len(paths))
    first = paths[:count]
    last = paths[-count:] if len(paths) > count else []
    combined: List[str] = []
    for entry in first + last:
        if entry not in combined:
            combined.append(entry)
    return combined


def _random_sample(paths: List[str], size: int, rng: random.Random) -> List[str]:
    if not paths or size <= 0:
        return []
    if size >= len(paths):
        return list(paths)
    return rng.sample(paths, size)


def _select_sample_specs(pairs: List[Tuple[str, Path, Path]], seed: int, top_bottom_only: bool) -> List[SheetSpec]:
    relative_paths = [relative for relative, _, _ in pairs]
    sample_40 = _first_last(relative_paths, 20)
    if top_bottom_only:
        return [SheetSpec("top_bottom_40", sample_40)]
    rng = random.Random(seed)
    sample_100 = _random_sample(relative_paths, 100, rng)
    sample_500 = _random_sample(relative_paths, 500, rng)
    return [
        SheetSpec("sample_40", sample_40),
        SheetSpec("sample_100", sample_100),
        SheetSpec("sample_500", sample_500),
    ]


def _collect_pending_paths(specs: List[SheetSpec], workbook: SampleWorkbook) -> List[str]:
    pending: List[str] = []
    seen: set[str] = set()
    for spec in specs:
        for relative in spec.paths:
            if workbook.has_entry(spec.name, relative):
                continue
            if relative in seen:
                continue
            pending.append(relative)
            seen.add(relative)
    return pending


def _run_sample_workbook(pipeline: MetricsPipeline, base_dir: Path, output_path: Path, seed: int, top_bottom_only: bool) -> None:
    pairs = pipeline.matched_pairs()
    if not pairs:
        raise RuntimeError("No matched input/output pairs were found under the base directory")
    specs = _select_sample_specs(pairs, seed, top_bottom_only)
    workbook = SampleWorkbook(output_path, specs, COLUMN_DEFINITIONS)
    pending = _collect_pending_paths(specs, workbook)
    records: Dict[str, Dict[str, object]] = {}
    if pending:
        progress_bar = tqdm(total=len(pending), unit="file", bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")
        def progress_callback(event: str, value) -> None:
            if event == "start":
                progress_bar.reset(total=int(value))
                progress_bar.refresh()
            elif event == "update":
                progress_bar.update(int(value))
        records = pipeline.compute_records_for_paths(pending, progress_callback=progress_callback)
        progress_bar.close()
    for spec in specs:
        for relative in spec.paths:
            if workbook.has_entry(spec.name, relative):
                continue
            record = records.get(relative)
            if not record:
                logging.warning("No metrics captured for %s", relative)
                continue
            workbook.append_record(spec.name, record)
    workbook.finalize(
        table_headers=TABLE_HEADERS,
        final_score_header="Final Score",
        decimal_headers=DECIMAL_HEADERS,
        integer_headers=INTEGER_HEADERS,
        thick_boundaries=THICK_BOUNDARIES,
    )
    logging.info("Wrote sample workbook to %s", output_path)


def _run_random_workbook(pipeline: MetricsPipeline, base_dir: Path, output_path: Path, count: int, seed: int) -> None:
    if count <= 0:
        raise ValueError("--random-sample must be positive")
    pairs = pipeline.matched_pairs()
    if not pairs:
        raise RuntimeError("No matched input/output pairs were found under the base directory")
    rng = random.Random(seed)
    if count >= len(pairs):
        selected = pairs
    else:
        selected = rng.sample(pairs, count)
    relative_paths = [relative for relative, _, _ in selected]
    spec = [SheetSpec(f"random_{count}", relative_paths)]
    workbook = SampleWorkbook(output_path, spec, COLUMN_DEFINITIONS)
    pending = _collect_pending_paths(spec, workbook)
    if not pending:
        logging.info("No pending paths for random sample, skipping workbook write")
        return
    progress_bar = tqdm(total=len(pending), unit="file", bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")
    def progress_callback(event: str, value) -> None:
        if event == "start":
            progress_bar.reset(total=int(value))
            progress_bar.refresh()
        elif event == "update":
            progress_bar.update(int(value))
    records = pipeline.compute_records_for_paths(pending, progress_callback=progress_callback)
    progress_bar.close()
    for relative in relative_paths:
        if workbook.has_entry(spec[0].name, relative):
            continue
        record = records.get(relative)
        if not record:
            logging.warning("No metrics captured for %s", relative)
            continue
        workbook.append_record(spec[0].name, record)
    workbook.finalize(
        table_headers=TABLE_HEADERS,
        final_score_header="Final Score",
        decimal_headers=DECIMAL_HEADERS,
        integer_headers=INTEGER_HEADERS,
        thick_boundaries=THICK_BOUNDARIES,
    )
    logging.info("Wrote random sample workbook (%d rows) to %s", len(relative_paths), output_path)


def main() -> None:
    args = parse_args()
    logging.basicConfig(level=getattr(logging, args.log_level.upper(), logging.INFO), format="%(message)s")
    base_dir = Path(args.base_dir).expanduser()
    input_dir = base_dir / "input"
    output_dir = base_dir / "output"
    if not input_dir.exists() or not output_dir.exists():
        raise FileNotFoundError("Base directory must contain input and output folders")
    device = require_cuda_device()
    logging.info("Using device: %s", device)
    output_path = _default_output_path(base_dir, args.output)
    resize_target = (512, 512) if args.resize_512 else None
    config = PipelineConfig(
        base_dir=base_dir,
        input_dir=input_dir,
        output_dir=output_dir,
        output_path=output_path,
        device=device,
        max_frames_per_video=args.max_frames_per_video,
        lpips_distance_max=args.lpips_distance_max,
        max_files=args.max_files,
        video_backend=args.video_backend,
        resize_to=resize_target,
    )
    pipeline = MetricsPipeline(config)
    if args.profile_only:
        result = pipeline.run_profile(sample_count=10, seed=args.seed, kind=args.profile_kind)
        _print_profile_summary(result)
        return
    if args.debug_random_10:
        records = pipeline.run_random_sample(seed=args.seed, image_count=5, video_count=5)
        if not records:
            raise RuntimeError("No files available for debug sampling")
        debug_output_path = Path(args.debug_output).expanduser()
        _write_debug_sample(records, debug_output_path)
        _print_debug_sample(records)
        logging.info("Wrote debug sample to %s", debug_output_path)
        return
    if args.random_sample:
        random_output = Path(args.random_output).expanduser() if args.random_output else (base_dir / f"random_sample_{args.random_sample}.xlsx")
        _run_random_workbook(pipeline, base_dir, random_output, args.random_sample, args.seed)
        return
    _run_sample_workbook(pipeline, base_dir, output_path, args.seed, args.top_bottom_40_only)


if __name__ == "__main__":
    main()



