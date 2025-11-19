from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class SheetSpec:
    name: str
    paths: List[str]


@dataclass(frozen=True)
class ColumnDefinition:
    header: str
    key: str


class SampleWorkbook:
    def __init__(self, path: Path, specs: Sequence[SheetSpec], columns: Sequence[ColumnDefinition]) -> None:
        self.path = path
        self.columns = list(columns)
        self.headers = [column.header for column in self.columns]
        self.specs = list(specs)
        if path.exists():
            self.workbook = load_workbook(path)
        else:
            self.workbook = Workbook()
            if self.specs:
                self.workbook.active.title = self.specs[0].name
        self.sheet_entries: Dict[str, set[str]] = {}
        for index, spec in enumerate(self.specs):
            sheet = self._ensure_sheet(spec.name, index == 0)
            self._ensure_headers(sheet)
            existing: set[str] = set()
            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                value = row[0]
                if value:
                    existing.add(str(value))
            self.sheet_entries[spec.name] = existing
        if not self.specs and "Sheet" in self.workbook.sheetnames:
            sheet = self.workbook["Sheet"]
            if sheet.max_row == 0:
                sheet.append(self.headers)
            self.sheet_entries[sheet.title] = set()

    def _ensure_sheet(self, name: str, first: bool) -> Worksheet:
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        if first and len(self.workbook.sheetnames) == 1 and self.workbook.active.max_row <= 1:
            sheet = self.workbook.active
            sheet.title = name
            return sheet
        return self.workbook.create_sheet(title=name)

    def _ensure_headers(self, sheet: Worksheet) -> None:
        if sheet.max_row == 0:
            sheet.append(self.headers)
            return
        first_row = [cell.value for cell in sheet[1][: len(self.headers)]]
        if first_row != self.headers:
            sheet.delete_rows(1, sheet.max_row)
            sheet.append(self.headers)

    def has_entry(self, sheet_name: str, filename: str) -> bool:
        entries = self.sheet_entries.get(sheet_name)
        if entries is None:
            return False
        return filename in entries

    def append_record(self, sheet_name: str, record: Dict[str, object]) -> None:
        sheet = self.workbook[sheet_name]
        row = [self._cell_value(record.get(column.key)) for column in self.columns]
        sheet.append(row)
        filename = str(record.get("filename", ""))
        if filename:
            self.sheet_entries[sheet_name].add(filename)
        self.save()

    def finalize(
        self,
        table_headers: Sequence[str],
        final_score_header: str,
        decimal_headers: Sequence[str],
        integer_headers: Sequence[str],
        thick_boundaries: Sequence[Tuple[str, str]],
    ) -> None:
        header_to_index = {column.header: index + 1 for index, column in enumerate(self.columns)}
        table_indices = [header_to_index[header] for header in table_headers if header in header_to_index]
        decimal_indices = {header_to_index[header] for header in decimal_headers if header in header_to_index}
        integer_indices = {header_to_index[header] for header in integer_headers if header in header_to_index}
        final_score_index = header_to_index.get(final_score_header, 0)
        boundary_pairs = [
            (header_to_index[left], header_to_index[right])
            for left, right in thick_boundaries
            if left in header_to_index and right in header_to_index
        ]
        for sheet_name in self.sheet_entries:
            sheet = self.workbook[sheet_name]
            self._format_sheet(
                sheet,
                table_indices,
                final_score_index,
                decimal_indices,
                integer_indices,
                boundary_pairs,
            )
        self.save()

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.path)

    def _cell_value(self, value: object) -> object:
        if value is None:
            return ""
        return value

    def _format_sheet(
        self,
        sheet: Worksheet,
        table_indices: Sequence[int],
        final_score_index: int,
        decimal_indices: set[int],
        integer_indices: set[int],
        boundary_pairs: Sequence[Tuple[int, int]],
    ) -> None:
        max_row = sheet.max_row
        if max_row <= 0:
            return
        alignment = Alignment(horizontal="center", vertical="center")
        for row in range(1, max_row + 1):
            for column in table_indices:
                sheet.cell(row=row, column=column).alignment = alignment
        for column in sorted(decimal_indices):
            for row in range(2, max_row + 1):
                cell = sheet.cell(row=row, column=column)
                if cell.value in ("", None):
                    continue
                cell.number_format = "0.0"
        for column in sorted(integer_indices):
            for row in range(2, max_row + 1):
                cell = sheet.cell(row=row, column=column)
                if cell.value in ("", None):
                    continue
                cell.number_format = "0"
        if final_score_index:
            for row in range(1, max_row + 1):
                cell = sheet.cell(row=row, column=final_score_index)
                existing = cell.font or Font()
                cell.font = Font(bold=True, name=existing.name, size=existing.sz)
        self._highlight_final_scores(sheet, final_score_index, max_row)
        self._apply_borders(sheet, table_indices, max_row, boundary_pairs)
        self._auto_fit_columns(sheet, table_indices, decimal_indices, integer_indices)

    def _apply_borders(
        self,
        sheet: Worksheet,
        table_indices: Sequence[int],
        max_row: int,
        boundary_pairs: Sequence[Tuple[int, int]],
    ) -> None:
        if not table_indices:
            return
        thick = Side(style="thick")
        sorted_indices = sorted(table_indices)
        first_col = sorted_indices[0]
        last_col = sorted_indices[-1]
        for column in sorted_indices:
            self._apply_border(sheet.cell(row=1, column=column), top=thick, bottom=thick)
            self._apply_border(sheet.cell(row=max_row, column=column), bottom=thick)
        for row in range(1, max_row + 1):
            self._apply_border(sheet.cell(row=row, column=first_col), left=thick)
            self._apply_border(sheet.cell(row=row, column=last_col), right=thick)
        for left_index, right_index in boundary_pairs:
            for row in range(1, max_row + 1):
                self._apply_border(sheet.cell(row=row, column=left_index), right=thick)
                self._apply_border(sheet.cell(row=row, column=right_index), left=thick)

    def _auto_fit_columns(
        self,
        sheet: Worksheet,
        table_indices: Sequence[int],
        decimal_indices: set[int],
        integer_indices: set[int],
    ) -> None:
        max_row = sheet.max_row
        decimal_set = set(decimal_indices)
        integer_set = set(integer_indices)
        for column_index in table_indices:
            header = self.columns[column_index - 1].header
            max_length = len(str(header))
            for row in range(2, max_row + 1):
                cell = sheet.cell(row=row, column=column_index)
                display = self._display_text(cell.value, column_index, decimal_set, integer_set)
                max_length = max(max_length, len(display))
            width = max(12, min(120, int(max_length * 1.2) + 2))
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    def _display_text(self, value: object, column_index: int, decimal_indices: set[int], integer_indices: set[int]) -> str:
        if value in (None, ""):
            return ""
        if column_index in decimal_indices:
            try:
                return f"{float(value):.1f}"
            except Exception:
                return str(value)
        if column_index in integer_indices:
            try:
                return f"{int(value)}"
            except Exception:
                return str(value)
        return str(value)

    def _apply_border(self, cell, left: Side | None = None, right: Side | None = None, top: Side | None = None, bottom: Side | None = None) -> None:
        cell.border = Border(
            left=left or cell.border.left,
            right=right or cell.border.right,
            top=top or cell.border.top,
            bottom=bottom or cell.border.bottom,
        )

    def _highlight_final_scores(self, sheet: Worksheet, column_index: int, max_row: int) -> None:
        if not column_index or max_row <= 1:
            return
        for row in range(2, max_row + 1):
            cell = sheet.cell(row=row, column=column_index)
            value = cell.value
            if value in ("", None):
                continue
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                continue
            if numeric >= 40.0:
                current_font = cell.font or Font()
                cell.font = Font(
                    bold=current_font.bold,
                    italic=current_font.italic,
                    name=current_font.name,
                    size=current_font.sz,
                    vertAlign=current_font.vertAlign,
                    underline=current_font.underline,
                    strike=current_font.strike,
                    color="9C0006",
                )
